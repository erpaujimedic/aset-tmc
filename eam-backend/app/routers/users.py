from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
import openpyxl
import io
from pydantic import BaseModel
from typing import List, Optional, Union
from app.database import supabase
from app.routers.auth import get_password_hash
import datetime

router = APIRouter(prefix="/users", tags=["Users"])

class UserCreate(BaseModel):
    name: str
    email: str
    username: str
    role: str
    branch: Union[str, List[str]]
    password: Optional[str] = None
    phone: Optional[str] = None
    status: Optional[str] = None
    department: Optional[str] = None
    workGroup: Optional[str] = None

class UserUpdate(BaseModel):
    name: str
    email: str
    username: str
    role: str
    branch: Union[str, List[str]]
    phone: Optional[str] = None
    status: Optional[str] = None
    department: Optional[str] = None
    workGroup: Optional[str] = None

class RoleCreate(BaseModel):
    role_name: str

@router.get("/roles")
def get_roles():
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    res = supabase.table("roles").select("name").execute()
    # Frontend expects an array of strings
    roles = [r["name"] for r in res.data]
    
    if "Public Access" not in roles:
        try:
            supabase.table("roles").insert({"name": "Public Access"}).execute()
            roles.append("Public Access")
        except:
            pass

    return {"data": roles}

@router.post("/roles")
def create_role(req: RoleCreate):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    supabase.table("roles").insert({"name": req.role_name}).execute()
    # Return updated list
    res = supabase.table("roles").select("name").execute()
    roles = [r["name"] for r in res.data]
    return {"data": roles}

@router.delete("/roles/{role_name}")
def delete_role(role_name: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    supabase.table("roles").delete().eq("name", role_name).execute()
    res = supabase.table("roles").select("name").execute()
    roles = [r["name"] for r in res.data]
    return {"data": roles}

@router.get("")
def get_users(branch: Optional[str] = None):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
        
    query = supabase.table("users").select("*")
    if branch and branch not in ("All Branches", "ALL", "ALL Branches"):
        query = query.eq("branch", branch)
    else:
        query = query.neq("branch", "DUMMY_SANDBOX")
        
    res = query.execute()
    
    # Map to frontend expected format
    formatted_users = []
    for u in res.data:
        formatted_users.append({
            "id": u["id"],
            "name": u["full_name"],
            "email": u["email"],
            "username": u["username"],
            "branch": u["branch"],
            "role": u["role"],
            "status": "Active", # default for now
            "lastLogin": "Never"
        })
    return {"data": formatted_users}

@router.post("")
def create_user(req: UserCreate):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    
    password = req.password if req.password else "admin123"
    hashed_password = get_password_hash(password)
    branch_val = req.branch[0] if isinstance(req.branch, list) and len(req.branch) > 0 else req.branch
    
    try:
        res = supabase.table("users").insert({
            "full_name": req.name,
            "email": req.email,
            "username": req.username,
            "role": req.role,
            "branch": branch_val if isinstance(branch_val, str) else str(branch_val),
            "password_hash": hashed_password
        }).execute()
        return {"data": res.data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/import-template")
def get_user_import_template():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Import Bulk Data
    ws1 = wb.active
    ws1.title = "Import Bulk Data"
    headers = [
        "Nama Lengkap (Wajib)", "Email (Wajib)", "Username (Wajib)", 
        "Role (Wajib)", "Cabang (Wajib)", "Password (Opsional)"
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    # Formatting widths
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 25
    ws1.column_dimensions["F"].width = 20
    
    ws1.freeze_panes = "A2"
    
    # Sheet 2: Information
    ws2 = wb.create_sheet(title="Information")
    
    # Fetch roles
    roles_res = supabase.table("roles").select("name").execute()
    roles = [r["name"] for r in roles_res.data] if roles_res.data else ["Master Admin", "Admin System", "Admin Checker", "User Biasa"]
    
    # Fetch branches
    branches_res = supabase.table("branches").select("name").execute()
    branches = [b["name"] for b in branches_res.data] if branches_res.data else ["TMC Pekanbaru", "Head Office"]
    
    info_data = [
        ["Informasi Pengisian Bulk Import User"],
        [],
        ["Keterangan Kolom:"],
        ["- Nama Lengkap", "Nama asli pengguna"],
        ["- Email", "Email pengguna (harus unik)"],
        ["- Username", "Username untuk login (harus unik)"],
        ["- Role", "Hak akses (Pilih dari daftar Role di bawah)"],
        ["- Cabang", "Cabang tugas (Pilih dari daftar Cabang di bawah atau 'All Branches')"],
        ["- Password", "Opsional. Jika dikosongkan, password otomatis menjadi 'admin123'"],
        [],
        ["DAFTAR ROLE YANG TERSEDIA:"]
    ]
    for r in roles:
        info_data.append(["", r])
        
    info_data.append([])
    info_data.append(["DAFTAR CABANG YANG TERSEDIA:"])
    info_data.append(["", "All Branches"])
    for b in branches:
        info_data.append(["", b])
        
    for row in info_data:
        ws2.append(row)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=User_Import_Template.xlsx"}
    )

@router.post("/import")
async def import_users(file: UploadFile = File(...)):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Hanya file .xlsx yang didukung")
        
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        if "Import Bulk Data" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Sheet 'Import Bulk Data' tidak ditemukan. Harap gunakan template yang benar.")
            
        ws = wb["Import Bulk Data"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="Tidak ada data yang ditemukan di sheet")
            
        users_to_insert = []
        for idx, row in enumerate(rows[1:], start=2):
            if not any(row): continue
            
            nama = str(row[0] or "").strip()
            email = str(row[1] or "").strip()
            username = str(row[2] or "").strip()
            role = str(row[3] or "").strip()
            cabang = str(row[4] or "").strip()
            password = str(row[5] or "").strip() if len(row) > 5 else ""
            
            if not nama or not email or not username or not role or not cabang:
                raise HTTPException(status_code=400, detail=f"Baris {idx}: Kolom wajib (Nama, Email, Username, Role, Cabang) tidak boleh kosong.")
                
            pwd_to_use = password if password and password != "None" else "admin123"
            hashed_password = get_password_hash(pwd_to_use)
            
            users_to_insert.append({
                "full_name": nama,
                "email": email,
                "username": username,
                "role": role,
                "branch": cabang,
                "password_hash": hashed_password
            })
            
        if not users_to_insert:
            raise HTTPException(status_code=400, detail="Tidak ada data valid untuk diimport")
            
        res = supabase.table("users").insert(users_to_insert).execute()
        return {"message": f"{len(users_to_insert)} user berhasil diimport", "data": res.data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/{user_id}")
def update_user(user_id: str, req: UserUpdate):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    branch_val = req.branch[0] if isinstance(req.branch, list) and len(req.branch) > 0 else req.branch
    
    try:
        res = supabase.table("users").update({
            "full_name": req.name,
            "email": req.email,
            "username": req.username,
            "role": req.role,
            "branch": branch_val if isinstance(branch_val, str) else str(branch_val)
        }).eq("id", user_id).execute()
        return {"data": res.data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/{user_id}")
def delete_user(user_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    supabase.table("users").delete().eq("id", user_id).execute()
    return {"message": "User deleted"}

@router.post("/{user_id}/reset-password")
def reset_user_password(user_id: str):
    if not supabase:
        raise HTTPException(status_code=500, detail="Database connection error")
    hashed_password = get_password_hash("admin123")
    supabase.table("users").update({"password_hash": hashed_password}).eq("id", user_id).execute()
    return {"message": "Password reset to admin123"}
